import os
from datetime import datetime
import openpyxl
from django.conf import settings
from rest_framework.serializers import Serializer
from rest_framework.response import Response
from django.contrib.auth import get_user_model
from django.db import transaction
from rest_framework import serializers, status
from .models import *
from .services.services import import_transaction
from djoser.serializers import (
	UserSerializer as BaseUserSerializer,
	UserCreateSerializer as BaseUserCreateSerializer, UsernameSerializer)

User = get_user_model()


class CustomUserCreateSerializer(BaseUserCreateSerializer):
	file = serializers.FileField(required=True, write_only=True)

	class Meta(BaseUserCreateSerializer.Meta):
		model = User
		fields = ['email', 'username', 'password', 'file']

	def validate(self, attrs):
		self._file = attrs.pop('file', None)
		return super().validate(attrs)

	def create(self, validated_data):
		user = super().create(validated_data)
		if self._file:
			import_result = self.import_transaction(user, self._file)
			user.import_result = import_result
		return user

	def to_representation(self, instance):
		data = super().to_representation(instance)
		if hasattr(instance, 'import_result'):
			data['import_result'] = instance.import_result
		return data

	def import_transaction(self, user, file):
		tags = {t['tag']: t['id'] for t in OperationTags.objects.values('id', 'tag')}
		row_count = ok_rows = 0
		try:
			if not file:
				raise FileNotFoundError('File not found, please try again')
			else:
				DataFile.objects.create(file=file, user=user, file_name=file.name)
			wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
			sheet = wb.active

			transactions_to_create = []
			f = DataFile.objects.values('id').get(file_name=file.name, user=user)
			for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=4, values_only=True):
				row_count += 1
				date, title, suma, tag = row
				if not any(row):
					continue

				if isinstance(date, str):
					transaction_date = datetime.strptime(date, '%Y-%m-%d').date()
				else:
					transaction_date = date.date()

				tag_id = tags.get(tag)
				if not tag_id:
					continue

				if len(title) > 40:
					continue
				else:
					title = title.strip()

				transaction_obj = UserInOutInfo(
					user=user,
					date=transaction_date,
					title=title,
					operation_type='income' if suma > 0 else 'expense',
					tag_id=tag_id,
					amount=abs(suma) if suma < 0 else suma,
					file_id=f.get('id'),
				)

				transactions_to_create.append(transaction_obj)

			if transactions_to_create:
				with transaction.atomic():
					UserInOutInfo.objects.bulk_create(transactions_to_create)
				ok_rows += 1

			if ok_rows == row_count:
				return {
					'status': 200,
					'message': f'Successfully import file, load {ok_rows} rows'
				}

			else:
				return {
					'status': 206,
					'message': f'Failed to import {row_count-ok_rows} rows, load {ok_rows} rows'
				}

		except Exception as e:
			User.objects.filter(username=user.username).delete()
			raise e


class CustomSetUsernameSerializer(UsernameSerializer):
	class Meta:
		model = User
		fields = ('username',)

	def validate(self, value):
		if User.objects.filter(username=value).exists():
			raise serializers.ValidationError("This username already exists.")
		return value


class CustomSetEmailSerializer(serializers.ModelSerializer):
	email = serializers.EmailField(required=True)

	class Meta:
		model = User
		fields = ('email',)

	def validate_email(self, value):
		if User.objects.filter(email=value).exclude(id=self.context['request'].user.id).exists():
			raise serializers.ValidationError("Email already exists.")
		return value

	def validate(self, attrs):
		attrs = super().validate(attrs)
		return attrs

