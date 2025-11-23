from datetime import datetime
import openpyxl
from django.db import transaction
from ..models import *


def import_transaction(user, file, reg):
	"""
	func which processes table files
	"""
	tags = {t['tag']: t['id'] for t in OperationTags.objects.values('id', 'tag')}
	row_count = ok_rows = 0
	f = None
	try:
		if not file:
			raise FileNotFoundError('File not found, please try again')
		else:
			f = DataFile.objects.create(file=file, user=user, file_name=file.name)
		wb = openpyxl.load_workbook(file.file, read_only=True, data_only=True)
		sheet = wb.active
		transactions_to_create = []
		for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=4, values_only=True):
			row_count += 1
			date, title, suma, tag = row

			if not any(row):
				continue

			if isinstance(date, str):
				transaction_date = datetime.strptime(date, '%Y-%m-%d').date()
			else:
				transaction_date = date.date()

			tag_id = tags.get(tag)
			if not tag_id:
				continue

			if not title or len(title) > 40:
				continue

			title = title.strip()

			transactions_to_create.append(
				UserInOutInfo(
					user=user,
					date=transaction_date,
					title=title,
					operation_type='income' if suma > 0 else 'expense',
					tag_id=tag_id,
					amount=abs(suma),
					file_id=f.id,
				)
			)

			ok_rows += 1

		if transactions_to_create:
			with transaction.atomic():
				UserInOutInfo.objects.bulk_create(transactions_to_create)

		if ok_rows == row_count:
			return {
				'status': 200,
				'message': f'Successfully import file, load {ok_rows} rows'
			}

		else:
			return {
				'status': 206,
				'message': f'Failed to import {row_count - ok_rows} rows, load {ok_rows} rows'
			}

	except Exception as e:
		if reg:
			User.objects.filter(username=user.username).delete()
		DataFile.objects.filter(id=f.id).delete()
		raise e